from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from mto_reader.models import ExtractedTable

# ---------------------------------------------------------------------------
# Shared style constants
# ---------------------------------------------------------------------------
_THIN = Side(style="thin")
_MEDIUM = Side(style="medium")

_BORDER_THIN = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_BORDER_MEDIUM = Border(left=_MEDIUM, right=_MEDIUM, top=_MEDIUM, bottom=_MEDIUM)

_TITLE_FILL = PatternFill(fill_type="solid", fgColor="BDD7EE")   # blue title bar
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9E1F2")  # lighter blue header

# Column widths keyed by a substring of the header text (lower-case)
_COL_WIDTHS: dict[str, float] = {
    "item": 6,
    "description": 52,
    "qty": 6,
    "wll": 9,
    "mbl": 9,
    "length": 10,
    "weight/unit": 18,
    "total weight": 15,
}


def _col_width(header: str) -> float:
    h = header.lower()
    for key, w in _COL_WIDTHS.items():
        if key in h:
            return w
    return max(len(header) + 4, 10)


def _is_description_col(header: str) -> bool:
    return "description" in header.lower()


def _border_row(ws, row: int, n_cols: int, border=_BORDER_THIN) -> None:
    """Apply a border to every cell in a row."""
    for col in range(1, n_cols + 1):
        ws.cell(row=row, column=col).border = border


def _write_rigging_block(
    ws,
    start_row: int,
    title: str,
    headers: list[str],
    data_rows: list[list],
    total_weight: str,
) -> int:
    """
    Write one complete rigging-table block (title + headers + data + footer)
    to *ws* beginning at *start_row*.  Returns the next available row number.

    Layout mirrors the PDF:
        Row 1  – Merged title bar (bold, blue, centred)
        Row 2  – Column headers  (bold, blue, centred, wrap)
        Row 3+ – Data rows       (centred; DESCRIPTION left-aligned)
        Last   – Total-weight footer (label right-aligned, value centred)
    """
    n_cols = len(headers)
    row = start_row

    # ------------------------------------------------------------------
    # Title row – spanning all columns
    # ------------------------------------------------------------------
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    title_cell = ws.cell(row=row, column=1)
    title_cell.value = title or "RIGGING MATERIAL LIST (OFFSHORE INSTALLATION)"
    title_cell.font = Font(bold=True, size=11)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = _TITLE_FILL
    title_cell.border = _BORDER_MEDIUM
    # Borders on the hidden cells of the merge (ensures outer box is drawn)
    for col in range(2, n_cols + 1):
        ws.cell(row=row, column=col).border = _BORDER_THIN
    ws.row_dimensions[row].height = 22
    row += 1

    # ------------------------------------------------------------------
    # Header row
    # ------------------------------------------------------------------
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = _HEADER_FILL
        cell.border = _BORDER_THIN
    ws.row_dimensions[row].height = 32
    row += 1

    # ------------------------------------------------------------------
    # Data rows
    # ------------------------------------------------------------------
    for data_row in data_rows:
        for col_idx, value in enumerate(data_row, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = value
            header = headers[col_idx - 1] if col_idx - 1 < len(headers) else ""
            if _is_description_col(header):
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _BORDER_THIN
        ws.row_dimensions[row].height = 15
        row += 1

    # ------------------------------------------------------------------
    # Total-weight footer row
    # ------------------------------------------------------------------
    if total_weight:
        label_col = n_cols - 1   # second-to-last column → "TOTAL WEIGHT :"
        value_col = n_cols        # last column           → numeric value
        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=col_idx)
            if col_idx == label_col:
                cell.value = "TOTAL WEIGHT :"
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx == value_col:
                cell.value = total_weight
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _BORDER_THIN
        ws.row_dimensions[row].height = 15
        row += 1

    return row  # first free row after this block


def export_tables_to_excel(
    tables: list[ExtractedTable],
    output_path: str,
    sheet_name: str = "MTO",
) -> None:
    if not tables:
        raise ValueError("No tables were extracted from the PDF.")

    rigging_tables = [t for t in tables if t.is_rigging_table]
    if not rigging_tables:
        raise ValueError("No Rigging Material List tables were found in the PDF.")

    destination = Path(output_path)
    destination.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_name or "Rigging MTO")[:31]

    current_row = 1
    for table in rigging_tables:
        df = table.dataframe
        headers = list(df.columns)
        data_rows = [list(row) for row in df.itertuples(index=False, name=None)]

        current_row = _write_rigging_block(
            ws=ws,
            start_row=current_row,
            title=table.table_title,
            headers=headers,
            data_rows=data_rows,
            total_weight=table.total_weight,
        )
        current_row += 1  # blank separator row between tables from different pages

    # Set column widths based on the headers from the first rigging table
    first_headers = list(rigging_tables[0].dataframe.columns)
    for col_idx, header in enumerate(first_headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = _col_width(header)

    # Freeze pane below the header row so column labels stay visible when scrolling
    ws.freeze_panes = ws.cell(row=3, column=1)

    wb.save(destination)

